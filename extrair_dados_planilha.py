#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from openpyxl import load_workbook
from datetime import datetime

# Carregar planilha
wb = load_workbook("/home/ubuntu/Controle_Residuos_Industriais.xlsx", data_only=True)

# ========================================
# EXTRAIR DADOS DA ABA LANÇAMENTOS
# ========================================
ws_lancamentos = wb["Lançamentos"]

lancamentos = []
for row in ws_lancamentos.iter_rows(min_row=2, max_row=100, values_only=True):
    # Pegar apenas as 6 primeiras colunas (A-F)
    row_data = row[:6]
    if len(row_data) < 6:
        continue
    codigo, descricao, quantidade, data, destinacao, observacoes = row_data
    
    # Parar quando encontrar linha vazia
    if not codigo:
        break
    
    # Converter data para string se existir
    if data:
        if isinstance(data, datetime):
            data_str = data.strftime("%Y-%m-%d")
        else:
            data_str = str(data)
    else:
        data_str = None
    
    lancamentos.append({
        "codigo": str(codigo) if codigo else "",
        "descricao": str(descricao) if descricao else "",
        "quantidade": float(quantidade) if quantidade else 0,
        "data": data_str,
        "destinacao": str(destinacao) if destinacao else "",
        "observacoes": str(observacoes) if observacoes else ""
    })

# ========================================
# EXTRAIR DADOS DA ABA RESUMO
# ========================================
ws_resumo = wb["Resumo por Código"]

resumo = []
for row in ws_resumo.iter_rows(min_row=2, max_row=50, values_only=True):
    # Pegar apenas as 4 primeiras colunas (A-D)
    row_data = row[:4]
    if len(row_data) < 4:
        continue
    codigo, descricao, quantidade_total, percentual = row_data
    
    # Parar quando encontrar linha vazia ou TOTAL GERAL
    if not codigo or codigo == "TOTAL GERAL":
        break
    
    resumo.append({
        "codigo": str(codigo) if codigo else "",
        "descricao": str(descricao) if descricao else "",
        "quantidade_total": float(quantidade_total) if quantidade_total else 0,
        "percentual": float(percentual) if percentual else 0
    })

# ========================================
# EXTRAIR DADOS DO CADASTRO
# ========================================
ws_cadastro = wb["Cadastro de Resíduos"]

cadastro = []
for row in ws_cadastro.iter_rows(min_row=2, max_row=100, values_only=True):
    codigo, descricao = row
    
    # Parar quando encontrar linha vazia
    if not codigo:
        break
    
    cadastro.append({
        "codigo": str(codigo) if codigo else "",
        "descricao": str(descricao) if descricao else ""
    })

# ========================================
# CALCULAR ESTATÍSTICAS
# ========================================

# Total de resíduos
total_residuos = sum(item["quantidade_total"] for item in resumo)

# Tipos de resíduos
tipos_residuos = len(resumo)

# Maior gerador
if resumo:
    maior_gerador = max(resumo, key=lambda x: x["quantidade_total"])
    maior_gerador_codigo = maior_gerador["codigo"]
    maior_gerador_quantidade = maior_gerador["quantidade_total"]
else:
    maior_gerador_codigo = "N/A"
    maior_gerador_quantidade = 0

# Total de lançamentos
total_lancamentos = len(lancamentos)

# Top 10
top10 = sorted(resumo, key=lambda x: x["quantidade_total"], reverse=True)[:10]

# Dados por destinação
destinacoes = {}
for lanc in lancamentos:
    dest = lanc["destinacao"]
    if dest:
        if dest not in destinacoes:
            destinacoes[dest] = 0
        destinacoes[dest] += lanc["quantidade"]

destinacoes_lista = [{"destinacao": k, "quantidade": v} for k, v in destinacoes.items()]

# ========================================
# CRIAR ESTRUTURA JSON
# ========================================

dados = {
    "kpis": {
        "total_residuos": round(total_residuos, 2),
        "tipos_residuos": tipos_residuos,
        "maior_gerador_codigo": maior_gerador_codigo,
        "maior_gerador_quantidade": round(maior_gerador_quantidade, 2),
        "total_lancamentos": total_lancamentos
    },
    "lancamentos": lancamentos,
    "resumo": resumo,
    "top10": top10,
    "cadastro": cadastro,
    "destinacoes": destinacoes_lista,
    "ultima_atualizacao": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
}

# Salvar JSON
with open("/home/ubuntu/dashboard/dados_residuos.json", "w", encoding="utf-8") as f:
    json.dump(dados, f, ensure_ascii=False, indent=2)

print("Dados extraídos com sucesso!")
print(f"- {len(lancamentos)} lançamentos")
print(f"- {len(resumo)} tipos de resíduos")
print(f"- {len(top10)} no Top 10")
print(f"- {len(destinacoes_lista)} tipos de destinação")
print(f"- Total: {total_residuos:.2f}")

