#!/usr/bin/env python3
"""
Script para atualizar planilha para versão 3.0
- Adiciona coluna "Unidade" nos lançamentos
- Cria aba "Cadastro de Destinatários"
- Cria aba "Cadastro de Transportadores"
- Adiciona campos de destinatário e transportador nos lançamentos
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def criar_estilo_cabecalho():
    """Retorna estilos para cabeçalho"""
    return {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def criar_estilo_celula():
    """Retorna estilos para células normais"""
    return {
        'font': Font(name='Calibri', size=11),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def atualizar_aba_lancamentos(wb):
    """Atualiza aba Lançamentos com coluna Unidade e novos campos"""
    
    print("📊 Atualizando aba 'Lançamentos'...")
    
    ws = wb['Lançamentos']
    
    # Salvar dados existentes
    dados_existentes = []
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value:  # Se tem código
            dados_existentes.append({
                'codigo': ws[f'A{row}'].value,
                'descricao': ws[f'B{row}'].value,
                'quantidade': ws[f'C{row}'].value,
                'data': ws[f'D{row}'].value,
                'destinacao': ws[f'E{row}'].value,
                'observacoes': ws[f'F{row}'].value
            })
    
    print(f"   📦 Dados existentes salvos: {len(dados_existentes)} lançamentos")
    
    # Limpar planilha
    ws.delete_rows(1, ws.max_row)
    
    # Novos cabeçalhos
    cabecalhos = [
        'Código do\nResíduo',
        'Descrição do\nResíduo',
        'Quantidade',
        'Unidade',
        'Data de\nGeração',
        'Destinação',
        'Destinatário',
        'Transportador',
        'Observações'
    ]
    
    # Escrever cabeçalhos
    estilo_cab = criar_estilo_cabecalho()
    for col, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=titulo)
        celula.font = estilo_cab['font']
        celula.fill = estilo_cab['fill']
        celula.alignment = estilo_cab['alignment']
        celula.border = estilo_cab['border']
    
    # Ajustar altura da linha de cabeçalho
    ws.row_dimensions[1].height = 30
    
    # Ajustar larguras das colunas
    larguras = {
        'A': 15,  # Código
        'B': 35,  # Descrição
        'C': 12,  # Quantidade
        'D': 10,  # Unidade
        'E': 12,  # Data
        'F': 15,  # Destinação
        'G': 30,  # Destinatário
        'H': 30,  # Transportador
        'I': 40   # Observações
    }
    
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    
    # Restaurar dados existentes
    estilo_cel = criar_estilo_celula()
    
    for idx, dados in enumerate(dados_existentes, 2):
        # Código
        ws[f'A{idx}'] = dados['codigo']
        
        # Descrição (fórmula PROCV)
        ws[f'B{idx}'] = f"=IFERROR(VLOOKUP(A{idx},'Cadastro de Resíduos'!A:B,2,FALSE),\"\")"
        
        # Quantidade
        ws[f'C{idx}'] = dados['quantidade']
        ws[f'C{idx}'].number_format = '#,##0.00'
        
        # Unidade (padrão: kg)
        ws[f'D{idx}'] = 'kg'
        
        # Data
        ws[f'E{idx}'] = dados['data']
        if isinstance(dados['data'], datetime):
            ws[f'E{idx}'].number_format = 'DD/MM/YYYY'
        
        # Destinação
        ws[f'F{idx}'] = dados['destinacao']
        
        # Destinatário (vazio por enquanto)
        ws[f'G{idx}'] = ''
        
        # Transportador (vazio por enquanto)
        ws[f'H{idx}'] = ''
        
        # Observações
        ws[f'I{idx}'] = dados['observacoes']
        
        # Aplicar estilos
        for col in 'ABCDEFGHI':
            celula = ws[f'{col}{idx}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
            if col in ['A', 'D', 'E', 'F']:
                celula.alignment = Alignment(horizontal='center', vertical='center')
            else:
                celula.alignment = estilo_cel['alignment']
    
    # Adicionar linhas vazias com fórmulas (até linha 1000)
    for row in range(len(dados_existentes) + 2, 1001):
        # Descrição (fórmula PROCV)
        ws[f'B{row}'] = f"=IFERROR(VLOOKUP(A{row},'Cadastro de Resíduos'!A:B,2,FALSE),\"\")"
        
        # Unidade padrão
        ws[f'D{row}'] = 'kg'
        
        # Aplicar estilos
        for col in 'ABCDEFGHI':
            celula = ws[f'{col}{row}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
    
    # Validação de dados - Código do Resíduo
    dv_codigo = DataValidation(
        type="list",
        formula1="='Cadastro de Resíduos'!$A$2:$A$1000",
        allow_blank=True
    )
    dv_codigo.error = 'Código inválido'
    dv_codigo.errorTitle = 'Código não cadastrado'
    dv_codigo.prompt = 'Selecione um código da lista'
    dv_codigo.promptTitle = 'Código do Resíduo'
    ws.add_data_validation(dv_codigo)
    dv_codigo.add(f'A2:A1000')
    
    # Validação de dados - Unidade
    dv_unidade = DataValidation(
        type="list",
        formula1='"kg,L,m³,ton,unidade"',
        allow_blank=False
    )
    dv_unidade.error = 'Unidade inválida'
    dv_unidade.errorTitle = 'Unidade não permitida'
    dv_unidade.prompt = 'Selecione: kg, L, m³, ton ou unidade'
    dv_unidade.promptTitle = 'Unidade de Medida'
    ws.add_data_validation(dv_unidade)
    dv_unidade.add(f'D2:D1000')
    
    # Validação de dados - Destinação
    dv_destinacao = DataValidation(
        type="list",
        formula1='"Reciclagem,Reuso,Aterro,Coprocessamento,Incineração,Tratamento,Compostagem,Outro"',
        allow_blank=True
    )
    dv_destinacao.error = 'Destinação inválida'
    dv_destinacao.errorTitle = 'Destinação não permitida'
    dv_destinacao.prompt = 'Selecione uma destinação da lista'
    dv_destinacao.promptTitle = 'Tipo de Destinação'
    ws.add_data_validation(dv_destinacao)
    dv_destinacao.add(f'F2:F1000')
    
    # Validação de dados - Destinatário
    dv_destinatario = DataValidation(
        type="list",
        formula1="='Cadastro de Destinatários'!$A$2:$A$1000",
        allow_blank=True
    )
    dv_destinatario.error = 'Destinatário inválido'
    dv_destinatario.errorTitle = 'Destinatário não cadastrado'
    dv_destinatario.prompt = 'Selecione um destinatário da lista'
    dv_destinatario.promptTitle = 'Destinatário'
    ws.add_data_validation(dv_destinatario)
    dv_destinatario.add(f'G2:G1000')
    
    # Validação de dados - Transportador
    dv_transportador = DataValidation(
        type="list",
        formula1="='Cadastro de Transportadores'!$A$2:$A$1000",
        allow_blank=True
    )
    dv_transportador.error = 'Transportador inválido'
    dv_transportador.errorTitle = 'Transportador não cadastrado'
    dv_transportador.prompt = 'Selecione um transportador da lista'
    dv_transportador.promptTitle = 'Transportador'
    ws.add_data_validation(dv_transportador)
    dv_transportador.add(f'H2:H1000')
    
    # Congelar painéis
    ws.freeze_panes = 'A2'
    
    # Ativar filtros
    ws.auto_filter.ref = f'A1:I1000'
    
    print("   ✅ Aba 'Lançamentos' atualizada com sucesso!")
    print(f"   ✅ {len(dados_existentes)} lançamentos restaurados")
    print("   ✅ Coluna 'Unidade' adicionada")
    print("   ✅ Colunas 'Destinatário' e 'Transportador' adicionadas")

def criar_aba_destinatarios(wb):
    """Cria aba Cadastro de Destinatários"""
    
    print("\n🏢 Criando aba 'Cadastro de Destinatários'...")
    
    # Criar nova aba
    if 'Cadastro de Destinatários' in wb.sheetnames:
        del wb['Cadastro de Destinatários']
    
    ws = wb.create_sheet('Cadastro de Destinatários')
    
    # Cabeçalhos
    cabecalhos = [
        'Nome da Empresa',
        'CNPJ',
        'Endereço',
        'CEP',
        'Cidade',
        'Estado',
        'Nº Licença de\nOperação',
        'Validade da\nLicença',
        'Telefone',
        'E-mail',
        'Observações'
    ]
    
    # Escrever cabeçalhos
    estilo_cab = criar_estilo_cabecalho()
    for col, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=titulo)
        celula.font = estilo_cab['font']
        celula.fill = estilo_cab['fill']
        celula.alignment = estilo_cab['alignment']
        celula.border = estilo_cab['border']
    
    # Ajustar altura da linha de cabeçalho
    ws.row_dimensions[1].height = 30
    
    # Ajustar larguras das colunas
    larguras = {
        'A': 35,  # Nome
        'B': 18,  # CNPJ
        'C': 40,  # Endereço
        'D': 12,  # CEP
        'E': 20,  # Cidade
        'F': 8,   # Estado
        'G': 20,  # Nº Licença
        'H': 15,  # Validade
        'I': 15,  # Telefone
        'J': 30,  # E-mail
        'K': 40   # Observações
    }
    
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    
    # Adicionar exemplos
    exemplos = [
        {
            'nome': 'EcoRecicla Ltda',
            'cnpj': '12.345.678/0001-90',
            'endereco': 'Rua das Indústrias, 1000',
            'cep': '01234-567',
            'cidade': 'São Paulo',
            'estado': 'SP',
            'licenca': 'LO-2024-001',
            'validade': datetime(2025, 12, 31),
            'telefone': '(11) 3456-7890',
            'email': 'contato@ecorecicla.com.br',
            'obs': 'Especializada em reciclagem de metais'
        },
        {
            'nome': 'Aterro Sanitário Central',
            'cnpj': '98.765.432/0001-10',
            'endereco': 'Rodovia BR-101, km 45',
            'cep': '98765-432',
            'cidade': 'Guarulhos',
            'estado': 'SP',
            'licenca': 'LO-2024-002',
            'validade': datetime(2026, 6, 30),
            'telefone': '(11) 2345-6789',
            'email': 'operacao@aterrocentral.com.br',
            'obs': 'Aterro classe II-A'
        },
        {
            'nome': 'Tratamento de Resíduos S/A',
            'cnpj': '11.222.333/0001-44',
            'endereco': 'Av. Industrial, 500',
            'cep': '13000-000',
            'cidade': 'Campinas',
            'estado': 'SP',
            'licenca': 'ISENTO',
            'validade': None,
            'telefone': '(19) 3333-4444',
            'email': 'comercial@tratamentoresiduos.com.br',
            'obs': 'Tratamento de resíduos químicos'
        }
    ]
    
    estilo_cel = criar_estilo_celula()
    
    for idx, exemplo in enumerate(exemplos, 2):
        ws[f'A{idx}'] = exemplo['nome']
        ws[f'B{idx}'] = exemplo['cnpj']
        ws[f'C{idx}'] = exemplo['endereco']
        ws[f'D{idx}'] = exemplo['cep']
        ws[f'E{idx}'] = exemplo['cidade']
        ws[f'F{idx}'] = exemplo['estado']
        ws[f'G{idx}'] = exemplo['licenca']
        ws[f'H{idx}'] = exemplo['validade']
        if exemplo['validade']:
            ws[f'H{idx}'].number_format = 'DD/MM/YYYY'
        ws[f'I{idx}'] = exemplo['telefone']
        ws[f'J{idx}'] = exemplo['email']
        ws[f'K{idx}'] = exemplo['obs']
        
        # Aplicar estilos
        for col in 'ABCDEFGHIJK':
            celula = ws[f'{col}{idx}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
            if col in ['B', 'D', 'F', 'G', 'H', 'I']:
                celula.alignment = Alignment(horizontal='center', vertical='center')
            else:
                celula.alignment = estilo_cel['alignment']
    
    # Adicionar linhas vazias com formatação (até linha 100)
    for row in range(5, 101):
        for col in 'ABCDEFGHIJK':
            celula = ws[f'{col}{row}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
    
    # Validação de dados - Estado
    dv_estado = DataValidation(
        type="list",
        formula1='"AC,AL,AP,AM,BA,CE,DF,ES,GO,MA,MT,MS,MG,PA,PB,PR,PE,PI,RJ,RN,RS,RO,RR,SC,SP,SE,TO"',
        allow_blank=True
    )
    dv_estado.error = 'Estado inválido'
    dv_estado.errorTitle = 'UF não encontrada'
    dv_estado.prompt = 'Selecione a UF do estado'
    dv_estado.promptTitle = 'Estado (UF)'
    ws.add_data_validation(dv_estado)
    dv_estado.add(f'F2:F100')
    
    # Congelar painéis
    ws.freeze_panes = 'A2'
    
    # Ativar filtros
    ws.auto_filter.ref = f'A1:K100'
    
    print("   ✅ Aba 'Cadastro de Destinatários' criada!")
    print(f"   ✅ {len(exemplos)} destinatários de exemplo adicionados")

def criar_aba_transportadores(wb):
    """Cria aba Cadastro de Transportadores"""
    
    print("\n🚛 Criando aba 'Cadastro de Transportadores'...")
    
    # Criar nova aba
    if 'Cadastro de Transportadores' in wb.sheetnames:
        del wb['Cadastro de Transportadores']
    
    ws = wb.create_sheet('Cadastro de Transportadores')
    
    # Cabeçalhos
    cabecalhos = [
        'Nome da Empresa',
        'CNPJ',
        'Endereço',
        'CEP',
        'Cidade',
        'Estado',
        'Nº Licença de\nOperação',
        'Validade da\nLicença',
        'Telefone',
        'E-mail',
        'Observações'
    ]
    
    # Escrever cabeçalhos
    estilo_cab = criar_estilo_cabecalho()
    for col, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=titulo)
        celula.font = estilo_cab['font']
        celula.fill = estilo_cab['fill']
        celula.alignment = estilo_cab['alignment']
        celula.border = estilo_cab['border']
    
    # Ajustar altura da linha de cabeçalho
    ws.row_dimensions[1].height = 30
    
    # Ajustar larguras das colunas
    larguras = {
        'A': 35,  # Nome
        'B': 18,  # CNPJ
        'C': 40,  # Endereço
        'D': 12,  # CEP
        'E': 20,  # Cidade
        'F': 8,   # Estado
        'G': 20,  # Nº Licença
        'H': 15,  # Validade
        'I': 15,  # Telefone
        'J': 30,  # E-mail
        'K': 40   # Observações
    }
    
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    
    # Adicionar exemplos
    exemplos = [
        {
            'nome': 'TransResíduos Transportes Ltda',
            'cnpj': '22.333.444/0001-55',
            'endereco': 'Av. Logística, 2000',
            'cep': '02000-000',
            'cidade': 'São Paulo',
            'estado': 'SP',
            'licenca': 'LT-2024-010',
            'validade': datetime(2025, 10, 15),
            'telefone': '(11) 4567-8901',
            'email': 'operacional@transresiduos.com.br',
            'obs': 'Transporte de resíduos classe I e II'
        },
        {
            'nome': 'Logística Ambiental Express',
            'cnpj': '33.444.555/0001-66',
            'endereco': 'Rua dos Transportes, 300',
            'cep': '13100-000',
            'cidade': 'Campinas',
            'estado': 'SP',
            'licenca': 'LT-2024-011',
            'validade': datetime(2026, 3, 20),
            'telefone': '(19) 2222-3333',
            'email': 'contato@logambiental.com.br',
            'obs': 'Frota com rastreamento GPS'
        },
        {
            'nome': 'EcoTrans Soluções Ambientais',
            'cnpj': '44.555.666/0001-77',
            'endereco': 'Rod. Anhanguera, km 100',
            'cep': '14000-000',
            'cidade': 'Jundiaí',
            'estado': 'SP',
            'licenca': 'ISENTO',
            'validade': None,
            'telefone': '(11) 5555-6666',
            'email': 'comercial@ecotrans.com.br',
            'obs': 'Especializada em resíduos hospitalares'
        }
    ]
    
    estilo_cel = criar_estilo_celula()
    
    for idx, exemplo in enumerate(exemplos, 2):
        ws[f'A{idx}'] = exemplo['nome']
        ws[f'B{idx}'] = exemplo['cnpj']
        ws[f'C{idx}'] = exemplo['endereco']
        ws[f'D{idx}'] = exemplo['cep']
        ws[f'E{idx}'] = exemplo['cidade']
        ws[f'F{idx}'] = exemplo['estado']
        ws[f'G{idx}'] = exemplo['licenca']
        ws[f'H{idx}'] = exemplo['validade']
        if exemplo['validade']:
            ws[f'H{idx}'].number_format = 'DD/MM/YYYY'
        ws[f'I{idx}'] = exemplo['telefone']
        ws[f'J{idx}'] = exemplo['email']
        ws[f'K{idx}'] = exemplo['obs']
        
        # Aplicar estilos
        for col in 'ABCDEFGHIJK':
            celula = ws[f'{col}{idx}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
            if col in ['B', 'D', 'F', 'G', 'H', 'I']:
                celula.alignment = Alignment(horizontal='center', vertical='center')
            else:
                celula.alignment = estilo_cel['alignment']
    
    # Adicionar linhas vazias com formatação (até linha 100)
    for row in range(5, 101):
        for col in 'ABCDEFGHIJK':
            celula = ws[f'{col}{row}']
            celula.font = estilo_cel['font']
            celula.border = estilo_cel['border']
    
    # Validação de dados - Estado
    dv_estado = DataValidation(
        type="list",
        formula1='"AC,AL,AP,AM,BA,CE,DF,ES,GO,MA,MT,MS,MG,PA,PB,PR,PE,PI,RJ,RN,RS,RO,RR,SC,SP,SE,TO"',
        allow_blank=True
    )
    dv_estado.error = 'Estado inválido'
    dv_estado.errorTitle = 'UF não encontrada'
    dv_estado.prompt = 'Selecione a UF do estado'
    dv_estado.promptTitle = 'Estado (UF)'
    ws.add_data_validation(dv_estado)
    dv_estado.add(f'F2:F100')
    
    # Congelar painéis
    ws.freeze_panes = 'A2'
    
    # Ativar filtros
    ws.auto_filter.ref = f'A1:K100'
    
    print("   ✅ Aba 'Cadastro de Transportadores' criada!")
    print(f"   ✅ {len(exemplos)} transportadores de exemplo adicionados")

def main():
    """Função principal"""
    
    print()
    print("=" * 70)
    print("🔄 ATUALIZAÇÃO PARA VERSÃO 3.0")
    print("=" * 70)
    print()
    print("Melhorias:")
    print("  ✅ Adicionar coluna 'Unidade' nos lançamentos")
    print("  ✅ Criar cadastro de Destinatários")
    print("  ✅ Criar cadastro de Transportadores")
    print("  ✅ Adicionar campos nos lançamentos")
    print()
    print("=" * 70)
    print()
    
    # Abrir planilha
    excel_file = 'Controle_Residuos_Industriais.xlsx'
    
    print(f"📂 Abrindo planilha: {excel_file}")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        print("   ✅ Planilha carregada com sucesso!")
        print()
        
        # Atualizar aba Lançamentos
        atualizar_aba_lancamentos(wb)
        
        # Criar aba Destinatários
        criar_aba_destinatarios(wb)
        
        # Criar aba Transportadores
        criar_aba_transportadores(wb)
        
        # Reordenar abas
        print("\n📑 Reordenando abas...")
        ordem_abas = [
            'Lançamentos',
            'Cadastro de Resíduos',
            'Cadastro de Destinatários',
            'Cadastro de Transportadores',
            'Resumo por Código'
        ]
        
        for aba in wb.sheetnames:
            if aba not in ordem_abas:
                ordem_abas.append(aba)
        
        for idx, nome_aba in enumerate(ordem_abas):
            if nome_aba in wb.sheetnames:
                wb.move_sheet(nome_aba, idx - len(wb.sheetnames))
        
        print("   ✅ Abas reordenadas!")
        
        # Salvar
        print("\n💾 Salvando planilha...")
        wb.save(excel_file)
        print(f"   ✅ Planilha salva: {excel_file}")
        
        print()
        print("=" * 70)
        print("✅ ATUALIZAÇÃO CONCLUÍDA COM SUCESSO!")
        print("=" * 70)
        print()
        print("📋 Resumo das alterações:")
        print()
        print("  📊 Aba 'Lançamentos':")
        print("     • Coluna 'Unidade' adicionada (kg, L, m³, ton, unidade)")
        print("     • Coluna 'Destinatário' adicionada")
        print("     • Coluna 'Transportador' adicionada")
        print("     • Validações de dados configuradas")
        print()
        print("  🏢 Aba 'Cadastro de Destinatários':")
        print("     • Nome, CNPJ, Endereço completo")
        print("     • Nº Licença de Operação (com opção ISENTO)")
        print("     • Validade da Licença")
        print("     • Telefone, E-mail, Observações")
        print("     • 3 exemplos pré-cadastrados")
        print()
        print("  🚛 Aba 'Cadastro de Transportadores':")
        print("     • Nome, CNPJ, Endereço completo")
        print("     • Nº Licença de Operação (com opção ISENTO)")
        print("     • Validade da Licença")
        print("     • Telefone, E-mail, Observações")
        print("     • 3 exemplos pré-cadastrados")
        print()
        print("💡 Próximos passos:")
        print("   1. Abrir a planilha Excel")
        print("   2. Verificar novas abas e campos")
        print("   3. Cadastrar seus destinatários e transportadores")
        print("   4. Fazer lançamentos com os novos campos")
        print("   5. Atualizar dashboard web (próximo passo)")
        print()
        print("=" * 70)
        
    except FileNotFoundError:
        print(f"❌ Erro: Arquivo {excel_file} não encontrado!")
        print("💡 Certifique-se de que o arquivo está no mesmo diretório")
    except Exception as e:
        print(f"❌ Erro durante atualização: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

