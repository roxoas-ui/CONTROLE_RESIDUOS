#!/usr/bin/env python3
"""
Script para adicionar múltiplos códigos de resíduos de uma vez
Facilita a adição em massa de novos códigos na planilha e no JSON
"""

import openpyxl
import json
import sys
import os

def adicionar_codigos_excel(excel_file, novos_codigos):
    """Adiciona códigos na planilha Excel"""
    
    print("=" * 60)
    print("📊 ADICIONANDO CÓDIGOS NA PLANILHA EXCEL")
    print("=" * 60)
    print()
    
    if not os.path.exists(excel_file):
        print(f"❌ Erro: Arquivo {excel_file} não encontrado!")
        return False
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        if 'Cadastro de Resíduos' not in wb.sheetnames:
            print(f"❌ Erro: Aba 'Cadastro de Resíduos' não encontrada!")
            return False
        
        ws = wb['Cadastro de Resíduos']
        
        # Encontrar última linha com dados
        ultima_linha = ws.max_row
        print(f"📍 Última linha atual: {ultima_linha}")
        print()
        
        # Verificar códigos duplicados
        codigos_existentes = set()
        for row in range(2, ultima_linha + 1):
            codigo = ws[f'A{row}'].value
            if codigo:
                codigos_existentes.add(str(codigo).strip().upper())
        
        print(f"📋 Códigos existentes: {len(codigos_existentes)}")
        print()
        
        # Adicionar novos códigos
        adicionados = 0
        duplicados = 0
        
        for codigo, descricao in novos_codigos:
            codigo_upper = str(codigo).strip().upper()
            
            if codigo_upper in codigos_existentes:
                print(f"⚠️  Código duplicado (ignorado): {codigo}")
                duplicados += 1
                continue
            
            linha = ultima_linha + adicionados + 1
            ws[f'A{linha}'] = codigo
            ws[f'B{linha}'] = descricao
            
            print(f"✅ Linha {linha}: {codigo} - {descricao}")
            
            codigos_existentes.add(codigo_upper)
            adicionados += 1
        
        print()
        
        # Salvar
        if adicionados > 0:
            wb.save(excel_file)
            print(f"💾 Planilha salva: {excel_file}")
            print(f"✅ Códigos adicionados: {adicionados}")
        else:
            print(f"⚠️  Nenhum código novo para adicionar")
        
        if duplicados > 0:
            print(f"⚠️  Códigos duplicados ignorados: {duplicados}")
        
        print()
        return adicionados > 0
        
    except Exception as e:
        print(f"❌ Erro ao processar planilha: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def adicionar_codigos_json(json_file, novos_codigos):
    """Adiciona códigos no arquivo JSON"""
    
    print("=" * 60)
    print("📄 ADICIONANDO CÓDIGOS NO ARQUIVO JSON")
    print("=" * 60)
    print()
    
    if not os.path.exists(json_file):
        print(f"⚠️  Arquivo {json_file} não encontrado")
        print(f"💡 Criando estrutura nova...")
        dados = {
            'kpis': {},
            'lancamentos': [],
            'resumo': [],
            'top10': [],
            'cadastro': [],
            'destinacoes': []
        }
    else:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                dados = json.load(f)
        except Exception as e:
            print(f"❌ Erro ao ler JSON: {str(e)}")
            return False
    
    # Verificar códigos existentes
    codigos_existentes = set()
    for item in dados.get('cadastro', []):
        codigo = item.get('codigo', '')
        if codigo:
            codigos_existentes.add(str(codigo).strip().upper())
    
    print(f"📋 Códigos existentes no JSON: {len(codigos_existentes)}")
    print()
    
    # Adicionar novos códigos
    adicionados = 0
    duplicados = 0
    
    for codigo, descricao in novos_codigos:
        codigo_upper = str(codigo).strip().upper()
        
        if codigo_upper in codigos_existentes:
            print(f"⚠️  Código duplicado (ignorado): {codigo}")
            duplicados += 1
            continue
        
        dados['cadastro'].append({
            'codigo': codigo,
            'descricao': descricao
        })
        
        print(f"✅ Adicionado: {codigo} - {descricao}")
        
        codigos_existentes.add(codigo_upper)
        adicionados += 1
    
    print()
    
    # Salvar
    if adicionados > 0:
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(dados, f, ensure_ascii=False, indent=2)
            
            print(f"💾 JSON salvo: {json_file}")
            print(f"✅ Códigos adicionados: {adicionados}")
        except Exception as e:
            print(f"❌ Erro ao salvar JSON: {str(e)}")
            return False
    else:
        print(f"⚠️  Nenhum código novo para adicionar")
    
    if duplicados > 0:
        print(f"⚠️  Códigos duplicados ignorados: {duplicados}")
    
    print()
    return adicionados > 0

def main():
    """Função principal"""
    
    print()
    print("=" * 60)
    print("🔧 FERRAMENTA DE ADIÇÃO DE CÓDIGOS DE RESÍDUOS")
    print("=" * 60)
    print()
    
    # ============================================================
    # CONFIGURAÇÃO: EDITE AQUI PARA ADICIONAR SEUS CÓDIGOS
    # ============================================================
    
    novos_codigos = [
        # Formato: ('CÓDIGO', 'Descrição do resíduo')
        
        # Exemplos:
        ('X100', 'Resíduo de borracha vulcanizada'),
        ('X101', 'Filtros de ar condicionado'),
        ('X102', 'Lâmpadas fluorescentes queimadas'),
        ('X103', 'Cartuchos de impressora vazios'),
        ('X104', 'Toner de impressora usado'),
        
        # Adicione mais códigos aqui:
        # ('X105', 'Seu resíduo aqui'),
        # ('X106', 'Outro resíduo'),
    ]
    
    # ============================================================
    # FIM DA CONFIGURAÇÃO
    # ============================================================
    
    if not novos_codigos:
        print("⚠️  Nenhum código para adicionar!")
        print("💡 Edite este script e adicione códigos na lista 'novos_codigos'")
        print()
        return
    
    print(f"📝 Códigos a adicionar: {len(novos_codigos)}")
    print()
    
    for codigo, descricao in novos_codigos:
        print(f"   • {codigo} - {descricao}")
    
    print()
    print("=" * 60)
    print()
    
    # Confirmar
    resposta = input("Deseja continuar? (s/n): ").strip().lower()
    if resposta != 's':
        print("❌ Operação cancelada pelo usuário")
        return
    
    print()
    
    # Adicionar na planilha
    excel_file = 'Controle_Residuos_Industriais.xlsx'
    sucesso_excel = adicionar_codigos_excel(excel_file, novos_codigos)
    
    print("=" * 60)
    print()
    
    # Adicionar no JSON
    json_file = 'dashboard/dados_residuos.json'
    sucesso_json = adicionar_codigos_json(json_file, novos_codigos)
    
    print("=" * 60)
    print()
    
    # Resumo final
    if sucesso_excel or sucesso_json:
        print("✅ PROCESSO CONCLUÍDO COM SUCESSO!")
        print()
        print("📋 Próximos passos:")
        print()
        
        if sucesso_excel:
            print("   1. Abrir a planilha Excel")
            print("   2. Verificar aba 'Cadastro de Resíduos'")
            print("   3. Testar na aba 'Lançamentos'")
            print()
        
        if sucesso_json:
            print("   4. Recarregar dashboard no navegador (F5)")
            print("   5. Clicar em 'Novo Lançamento'")
            print("   6. Verificar se novos códigos aparecem")
            print()
        
        print("   7. Se dashboard está online, fazer deploy:")
        print("      cd dashboard")
        print("      git add .")
        print("      git commit -m 'Novos códigos de resíduos'")
        print("      git push")
        print()
    else:
        print("⚠️  Nenhuma alteração foi feita")
        print()
    
    print("=" * 60)

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n❌ Operação cancelada pelo usuário (Ctrl+C)")
    except Exception as e:
        print(f"\n❌ Erro inesperado: {str(e)}")
        import traceback
        traceback.print_exc()

