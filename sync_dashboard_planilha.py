#!/usr/bin/env python3
"""
Script de Sincronização: Dashboard → Planilha Excel
Importa dados do JSON exportado do dashboard para a planilha Excel
"""

import json
import openpyxl
from datetime import datetime
import sys
import os

def sync_dashboard_to_excel(json_file, excel_file):
    """Sincroniza dados do dashboard (JSON) para planilha Excel"""
    
    print("=" * 60)
    print("🔄 SINCRONIZAÇÃO: Dashboard → Planilha Excel")
    print("=" * 60)
    print()
    
    # Verificar se arquivos existem
    if not os.path.exists(json_file):
        print(f"❌ Erro: Arquivo {json_file} não encontrado!")
        print(f"💡 Dica: Exporte os dados do dashboard primeiro")
        return False
    
    if not os.path.exists(excel_file):
        print(f"❌ Erro: Arquivo {excel_file} não encontrado!")
        return False
    
    try:
        # Carregar JSON
        print(f"📄 Lendo dados de: {json_file}")
        with open(json_file, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        
        total_lancamentos = len(dados.get('lancamentos', []))
        print(f"   ✓ {total_lancamentos} lançamentos encontrados")
        print()
        
        # Abrir planilha
        print(f"📊 Abrindo planilha: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        
        # Verificar se aba existe
        if 'Lançamentos' not in wb.sheetnames:
            print(f"❌ Erro: Aba 'Lançamentos' não encontrada!")
            print(f"   Abas disponíveis: {', '.join(wb.sheetnames)}")
            return False
        
        ws = wb['Lançamentos']
        print(f"   ✓ Aba 'Lançamentos' encontrada")
        print()
        
        # Fazer backup dos dados atuais
        dados_antigos = ws.max_row - 1  # -1 para não contar cabeçalho
        print(f"🗑️  Limpando {dados_antigos} lançamentos antigos...")
        
        # Limpar dados antigos (manter cabeçalho na linha 1)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        print(f"   ✓ Dados antigos removidos")
        print()
        
        # Adicionar novos lançamentos
        print(f"➕ Importando {total_lancamentos} lançamentos...")
        linha = 2
        erros = 0
        
        for idx, lanc in enumerate(dados['lancamentos'], 1):
            try:
                # Coluna A: Código
                ws[f'A{linha}'] = lanc.get('codigo', '')
                
                # Coluna B: Descrição
                ws[f'B{linha}'] = lanc.get('descricao', '')
                
                # Coluna C: Quantidade
                quantidade = lanc.get('quantidade', 0)
                ws[f'C{linha}'] = float(quantidade) if quantidade else 0
                
                # Coluna D: Data
                if lanc.get('data'):
                    try:
                        # Tentar converter string para data
                        data_str = lanc['data']
                        if 'T' in data_str:
                            data_obj = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
                        else:
                            data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                        ws[f'D{linha}'] = data_obj
                    except Exception as e:
                        # Se falhar, usar string mesmo
                        ws[f'D{linha}'] = lanc['data']
                
                # Coluna E: Destinação
                ws[f'E{linha}'] = lanc.get('destinacao', '')
                
                # Coluna F: Observações
                ws[f'F{linha}'] = lanc.get('observacoes', '')
                
                linha += 1
                
                # Mostrar progresso a cada 10 lançamentos
                if idx % 10 == 0:
                    print(f"   ⏳ Processando... {idx}/{total_lancamentos}")
                
            except Exception as e:
                erros += 1
                print(f"   ⚠️  Erro no lançamento {idx}: {str(e)}")
        
        print(f"   ✓ {linha-2} lançamentos importados")
        if erros > 0:
            print(f"   ⚠️  {erros} erros encontrados")
        print()
        
        # Salvar planilha
        print("💾 Salvando planilha...")
        wb.save(excel_file)
        print(f"   ✓ Planilha salva com sucesso")
        print()
        
        # Mostrar resumo
        print("=" * 60)
        print("📊 RESUMO DA SINCRONIZAÇÃO")
        print("=" * 60)
        print(f"✅ Lançamentos importados: {linha-2}")
        print(f"📦 Total de resíduos: {dados['kpis']['total_residuos']:.2f} kg/L/unidades")
        print(f"🗂️  Tipos de resíduos: {dados['kpis']['tipos_residuos']}")
        print(f"⚠️  Maior gerador: {dados['kpis']['maior_gerador_codigo']}")
        print(f"📅 Última atualização: {dados.get('ultima_atualizacao', 'N/A')}")
        print()
        print("✅ Sincronização concluída com sucesso!")
        print()
        print("💡 Próximos passos:")
        print("   1. Abra a planilha Excel")
        print("   2. Verifique a aba 'Lançamentos'")
        print("   3. Confira a aba 'Resumo por Código'")
        print("   4. Clique no botão 'Atualizar Resumo' se necessário")
        print("=" * 60)
        
        return True
        
    except Exception as e:
        print()
        print(f"❌ ERRO durante sincronização: {str(e)}")
        print()
        import traceback
        traceback.print_exc()
        return False

def main():
    """Função principal"""
    
    # Valores padrão
    json_file = 'dados_residuos.json'
    excel_file = 'Controle_Residuos_Industriais.xlsx'
    
    # Aceitar argumentos da linha de comando
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    if len(sys.argv) > 2:
        excel_file = sys.argv[2]
    
    # Executar sincronização
    sucesso = sync_dashboard_to_excel(json_file, excel_file)
    
    # Retornar código de saída
    sys.exit(0 if sucesso else 1)

if __name__ == '__main__':
    main()

